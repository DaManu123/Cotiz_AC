import os
from datetime import datetime
from flask import current_app
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def calcular_totales(items, descuento_tipo, descuento_valor, iva_porcentaje, envio):
    """
    Calcula todos los totales de una cotización.
    
    Args:
        items: lista de diccionarios con 'cantidad' y 'precio_unitario'
        descuento_tipo: 'percent' o 'fixed'
        descuento_valor: valor del descuento
        iva_porcentaje: porcentaje de IVA
        envio: monto de envío
    
    Returns:
        dict con subtotal, descuento_monto, base_impuesto, iva_monto, total
    """
    # Calcular subtotal
    subtotal = 0.0
    for item in items:
        total_item = round(item['cantidad'] * item['precio_unitario'], 2)
        subtotal += total_item
    subtotal = round(subtotal, 2)
    
    # Calcular descuento
    if descuento_tipo == 'percent':
        descuento_monto = round(subtotal * (descuento_valor / 100), 2)
    else:
        descuento_monto = round(descuento_valor, 2)
    
    # Base imponible
    base_impuesto = round(subtotal - descuento_monto, 2)
    
    # IVA
    iva_monto = round(base_impuesto * (iva_porcentaje / 100), 2)
    
    # Total
    total = round(base_impuesto + iva_monto + envio, 2)
    
    return {
        'subtotal': subtotal,
        'descuento_monto': descuento_monto,
        'base_impuesto': base_impuesto,
        'iva_monto': iva_monto,
        'total': total
    }

def exportar_excel(cotizacion, path):
    """
    Exporta una cotización a formato Excel con diseño profesional.
    """
    from app.models import Item
    
    # Crear DataFrame para los items
    items = Item.query.filter_by(cotizacion_id=cotizacion.id).all()
    
    data = []
    for item in items:
        data.append({
            'Cantidad': item.cantidad,
            'Unidad': item.unidad,
            'Descripción': item.descripcion,
            'P.U.': item.precio_unitario,
            'Total': item.total_item
        })
    
    df = pd.DataFrame(data)
    
    # Crear archivo Excel
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # Escribir datos en una posición específica
        df.to_excel(writer, sheet_name=f'Cotizacion_{cotizacion.id}', startrow=10, index=False)
        
        workbook = writer.book
        worksheet = writer.sheets[f'Cotizacion_{cotizacion.id}']
        
        # Estilos
        header_font = Font(name='Arial', size=12, bold=True)
        title_font = Font(name='Arial', size=14, bold=True)
        normal_font = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Encabezado empresa
        worksheet['A1'] = current_app.config['EMPRESA_NOMBRE']
        worksheet['A1'].font = title_font
        worksheet['A2'] = f"RFC: {current_app.config['EMPRESA_RFC']}"
        worksheet['A3'] = current_app.config['EMPRESA_DIRECCION']
        worksheet['A4'] = f"Tel: {current_app.config['EMPRESA_TELEFONO']}"
        worksheet['A5'] = f"Email: {current_app.config['EMPRESA_EMAIL']}"
        
        # Información de cotización
        worksheet['E1'] = 'COTIZACIÓN'
        worksheet['E1'].font = title_font
        worksheet['E2'] = f"No. {cotizacion.id}"
        worksheet['E3'] = f"Fecha: {cotizacion.fecha.strftime('%d/%m/%Y')}"
        
        # Información del cliente
        if cotizacion.cliente:
            worksheet['A7'] = 'CLIENTE:'
            worksheet['A7'].font = header_font
            worksheet['B7'] = cotizacion.cliente.nombre or ''
            worksheet['A8'] = 'Teléfono:'
            worksheet['B8'] = cotizacion.cliente.telefono or ''
            worksheet['A9'] = 'Correo:'
            worksheet['B9'] = cotizacion.cliente.correo or ''
        
        # Formatear encabezados de tabla
        for col in range(1, 6):
            cell = worksheet.cell(row=11, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formatear datos de items
        for row in range(12, 12 + len(items)):
            for col in range(1, 6):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                cell.font = normal_font
                
                # Formato moneda para precios
                if col in [4, 5]:  # P.U. y Total
                    cell.number_format = '"$"#,##0.00'
        
        # Resumen financiero
        resumen_row = 12 + len(items) + 2
        
        worksheet[f'D{resumen_row}'] = 'Subtotal:'
        worksheet[f'D{resumen_row}'].font = header_font
        worksheet[f'E{resumen_row}'] = cotizacion.subtotal
        worksheet[f'E{resumen_row}'].number_format = '"$"#,##0.00'
        
        resumen_row += 1
        desc_text = f"Descuento ({cotizacion.descuento_valor}{'%' if cotizacion.descuento_tipo == 'percent' else ' MXN'}):"
        worksheet[f'D{resumen_row}'] = desc_text
        worksheet[f'D{resumen_row}'].font = header_font
        worksheet[f'E{resumen_row}'] = -cotizacion.descuento_monto
        worksheet[f'E{resumen_row}'].number_format = '"$"#,##0.00'
        
        resumen_row += 1
        worksheet[f'D{resumen_row}'] = 'Base:'
        worksheet[f'D{resumen_row}'].font = header_font
        base = cotizacion.subtotal - cotizacion.descuento_monto
        worksheet[f'E{resumen_row}'] = base
        worksheet[f'E{resumen_row}'].number_format = '"$"#,##0.00'
        
        resumen_row += 1
        worksheet[f'D{resumen_row}'] = f'IVA ({cotizacion.iva_porcentaje}%):'
        worksheet[f'D{resumen_row}'].font = header_font
        worksheet[f'E{resumen_row}'] = cotizacion.iva_monto
        worksheet[f'E{resumen_row}'].number_format = '"$"#,##0.00'
        
        if cotizacion.envio > 0:
            resumen_row += 1
            worksheet[f'D{resumen_row}'] = 'Envío:'
            worksheet[f'D{resumen_row}'].font = header_font
            worksheet[f'E{resumen_row}'] = cotizacion.envio
            worksheet[f'E{resumen_row}'].number_format = '"$"#,##0.00'
        
        resumen_row += 1
        worksheet[f'D{resumen_row}'] = 'TOTAL:'
        worksheet[f'D{resumen_row}'].font = Font(name='Arial', size=12, bold=True)
        worksheet[f'E{resumen_row}'] = cotizacion.total
        worksheet[f'E{resumen_row}'].number_format = '"$"#,##0.00'
        worksheet[f'E{resumen_row}'].font = Font(name='Arial', size=12, bold=True)
        
        # Términos y condiciones
        terms_row = resumen_row + 3
        worksheet[f'A{terms_row}'] = 'TÉRMINOS Y CONDICIONES'
        worksheet[f'A{terms_row}'].font = header_font
        
        terms_text = cotizacion.terms_conditions or current_app.config['TERMINOS_CONDICIONES']
        terms_row += 1
        worksheet[f'A{terms_row}'] = terms_text
        worksheet[f'A{terms_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        worksheet.merge_cells(f'A{terms_row}:E{terms_row + 10}')
        
        # Firma
        firma_row = terms_row + 12
        worksheet[f'A{firma_row}'] = '_' * 40
        worksheet[f'A{firma_row + 1}'] = 'Firma y sello'
        worksheet[f'A{firma_row + 1}'].alignment = Alignment(horizontal='center')
        
        # Ajustar anchos de columna
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 10
        worksheet.column_dimensions['C'].width = 50
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15
    
    return path

def generar_pdf(cotizacion, path):
    """
    Genera un PDF de la cotización usando WeasyPrint.
    """
    from flask import render_template
    
    try:
        from weasyprint import HTML, CSS
        use_weasyprint = True
    except ImportError:
        use_weasyprint = False
    
    # Renderizar HTML
    html_content = render_template('cotizacion_printable.html', cotizacion=cotizacion)
    
    if use_weasyprint:
        # Usar WeasyPrint
        css_path = os.path.join(current_app.root_path, 'static', 'css', 'print.css')
        css = CSS(filename=css_path)
        HTML(string=html_content).write_pdf(path, stylesheets=[css])
    else:
        # Fallback a ReportLab
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        
        doc = SimpleDocTemplate(path, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph(f"COTIZACIÓN No. {cotizacion.id}", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Información empresa y cliente
        info_data = [
            [current_app.config['EMPRESA_NOMBRE'], f"Fecha: {cotizacion.fecha.strftime('%d/%m/%Y')}"],
            [f"RFC: {current_app.config['EMPRESA_RFC']}", ''],
            [current_app.config['EMPRESA_TELEFONO'], '']
        ]
        
        if cotizacion.cliente:
            info_data.append(['', ''])
            info_data.append([f"Cliente: {cotizacion.cliente.nombre or ''}", ''])
            if cotizacion.cliente.telefono:
                info_data.append([f"Tel: {cotizacion.cliente.telefono}", ''])
        
        info_table = Table(info_data, colWidths=[4*inch, 2*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabla de items
        items_data = [['Cant.', 'Unidad', 'Descripción', 'P.U.', 'Total']]
        for item in cotizacion.items:
            items_data.append([
                str(item.cantidad),
                item.unidad,
                item.descripcion[:60],
                f"${item.precio_unitario:,.2f}",
                f"${item.total_item:,.2f}"
            ])
        
        items_table = Table(items_data, colWidths=[0.6*inch, 0.8*inch, 3*inch, 1*inch, 1*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Resumen
        resumen_data = [
            ['Subtotal:', f"${cotizacion.subtotal:,.2f}"],
            [f"Descuento ({cotizacion.descuento_valor}{'%' if cotizacion.descuento_tipo == 'percent' else ' MXN'}):", f"${cotizacion.descuento_monto:,.2f}"],
            [f"IVA ({cotizacion.iva_porcentaje}%):", f"${cotizacion.iva_monto:,.2f}"],
        ]
        
        if cotizacion.envio > 0:
            resumen_data.append(['Envío:', f"${cotizacion.envio:,.2f}"])
        
        resumen_data.append(['TOTAL:', f"${cotizacion.total:,.2f}"])
        
        resumen_table = Table(resumen_data, colWidths=[4.5*inch, 1.5*inch])
        resumen_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -2), 'Helvetica'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ]))
        elements.append(resumen_table)
        
        doc.build(elements)
    
    return path
