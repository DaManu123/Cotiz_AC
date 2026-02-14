import os
from datetime import datetime
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XlImage  # type: ignore


class ExcelService:
    """
    Servicio para generación de archivos Excel de cotizaciones.
    Replica exactamente el formato Pro-Forma de Multiservicios RMG.

    Estructura de columnas (5 visuales, tamaño carta vertical):
        A=14  (IVA)
        B=9   (CANT.)
        C=42  (DESCRIPCIÓN)
        D=16  (P. UNITARIO)
        E=18  (TOTAL)
    """

    # ── Colores corporativos ──
    AZUL_CORP = '08568D'
    GRIS_CORP = 'F3F3F3'
    BLANCO = 'FFFFFF'
    NEGRO = '000000'

    # ── Anchos de columna ──
    COL_WIDTHS = {'A': 14, 'B': 9, 'C': 42, 'D': 16, 'E': 18}

    # ── Ruta del logo ──
    LOGO_PATH = os.path.join('static', 'img', 'logormg.jpg')

    def __init__(self, output_dir: str = 'exports/excel'):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    # ══════════════════════════════════════════════════════════
    #  ESTILOS
    # ══════════════════════════════════════════════════════════

    def _crear_estilos(self) -> Dict[str, Any]:
        fill_azul = PatternFill(start_color=self.AZUL_CORP, end_color=self.AZUL_CORP, fill_type='solid')
        fill_gris = PatternFill(start_color=self.GRIS_CORP, end_color=self.GRIS_CORP, fill_type='solid')
        fill_blanco = PatternFill(start_color=self.BLANCO, end_color=self.BLANCO, fill_type='solid')

        border_thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )
        border_bottom_azul = Border(bottom=Side(style='medium', color=self.AZUL_CORP))
        border_box_thin = Border(
            left=Side(style='thin', color=self.AZUL_CORP),
            right=Side(style='thin', color=self.AZUL_CORP),
            top=Side(style='thin', color=self.AZUL_CORP),
            bottom=Side(style='thin', color=self.AZUL_CORP),
        )

        return {
            'fill_azul': fill_azul,
            'fill_gris': fill_gris,
            'fill_blanco': fill_blanco,
            'border_thin': border_thin,
            'border_bottom_azul': border_bottom_azul,
            'border_box': border_box_thin,
            # Fuentes
            'font_proforma': Font(name='Arial', size=22, bold=True, color=self.AZUL_CORP),
            'font_empresa_dato': Font(name='Arial', size=9, color='555555'),
            'font_empresa_dato_bold': Font(name='Arial', size=9, bold=True, color='555555'),
            'font_info_label': Font(name='Arial', size=10, bold=True, color=self.AZUL_CORP),
            'font_info_value': Font(name='Arial', size=10, color=self.NEGRO),
            'font_header': Font(name='Arial', size=11, bold=True, color=self.BLANCO),
            'font_data': Font(name='Arial', size=10, color=self.NEGRO),
            'font_grupo': Font(name='Arial', size=10, bold=True, color=self.AZUL_CORP),
            'font_totales_label': Font(name='Arial', size=10, bold=True, color=self.NEGRO),
            'font_totales_value': Font(name='Arial', size=10, color=self.NEGRO),
            'font_pagado_label': Font(name='Arial', size=11, bold=True, color=self.BLANCO),
            'font_pagado_value': Font(name='Arial', size=11, bold=True, color=self.BLANCO),
            'font_terms_title': Font(name='Arial', size=9, bold=True, color=self.AZUL_CORP),
            'font_terms': Font(name='Arial', size=8, color='555555'),
            'font_footer': Font(name='Arial', size=9, bold=True, color=self.AZUL_CORP),
            'font_footer_sub': Font(name='Arial', size=8, italic=True, color='888888'),
            # Alineaciones
            'align_center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'align_left': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'align_right': Alignment(horizontal='right', vertical='center'),
        }

    # ══════════════════════════════════════════════════════════
    #  CONFIGURACIÓN DE HOJA
    # ══════════════════════════════════════════════════════════

    def _configurar_hoja(self, ws: Worksheet) -> None:
        ws.title = "Pro-Forma"
        for col_letter, width in self.COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER  # type: ignore
        ws.page_setup.orientation = 'portrait'
        ws.sheet_properties.pageSetUpPr.fitToPage = True  # type: ignore
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.3
        ws.page_margins.bottom = 0.3
        ws.sheet_view.showGridLines = False  # type: ignore

    # ══════════════════════════════════════════════════════════
    #  ENCABEZADO: LOGO + PRO-FORMA + INFO EMPRESA + FECHA/N°
    # ══════════════════════════════════════════════════════════

    def _escribir_encabezado(
        self, ws: Worksheet, empresa_data: Dict[str, Any],
        cotizacion_data: Dict[str, Any], estilos: Dict[str, Any]
    ) -> int:
        """
        Rows 1-10 approximately:
        - Logo image in A1 area (spanning ~5 rows, cols A-B)
        - "PRO-FORMA" large text in D1:E1
        - Company info (address, phone, email, fb, RFC) in A6:C10
        - Fecha / N° de Pro-forma boxes in D5:E6
        Returns next row.
        """
        s = estilos

        # ── Logo ──
        if os.path.exists(self.LOGO_PATH):
            img = XlImage(self.LOGO_PATH)
            img.width = 180
            img.height = 80
            ws.add_image(img, 'A1')

        # ── PRO-FORMA text ──
        ws.merge_cells('D1:E3')
        cell_pf = ws['D1']
        cell_pf.value = "PRO-FORMA"  # type: ignore
        cell_pf.font = s['font_proforma']
        cell_pf.alignment = Alignment(horizontal='right', vertical='center')

        # Adjust row heights for header area
        for r in range(1, 5):
            ws.row_dimensions[r].height = 18

        # ── Company info lines (rows 5-9, cols A:C) ──
        info_lines: List[str] = []
        if empresa_data.get('direccion'):
            info_lines.append(f"📍  {empresa_data['direccion']}")
        if empresa_data.get('telefono'):
            info_lines.append(f"📞  {empresa_data['telefono']}")
        if empresa_data.get('email'):
            info_lines.append(f"✉️  {empresa_data['email']}")
        if empresa_data.get('redes_sociales'):
            info_lines.append(f"🌐  {empresa_data['redes_sociales']}")
        if empresa_data.get('rfc'):
            info_lines.append(f"🆔  {empresa_data['rfc']}")

        row = 5
        for info in info_lines:
            ws.merge_cells(f'A{row}:C{row}')
            c = ws.cell(row=row, column=1)
            c.value = info  # type: ignore
            c.font = s['font_empresa_dato']
            c.alignment = s['align_left']
            ws.row_dimensions[row].height = 15
            row += 1

        # ── Fecha / N° de Pro-forma boxes (D5:E6) ──
        fecha_str = cotizacion_data.get('fecha', '')
        try:
            fecha_obj = datetime.strptime(str(fecha_str), '%Y-%m-%d')
            fecha_fmt = fecha_obj.strftime('%d/%m/%Y')
        except (ValueError, TypeError):
            fecha_fmt = str(fecha_str) if fecha_str else datetime.now().strftime('%d/%m/%Y')

        # Fecha label D5
        c_fl = ws.cell(row=5, column=4)
        c_fl.value = "Fecha"  # type: ignore
        c_fl.font = s['font_info_label']
        c_fl.alignment = s['align_center']
        c_fl.border = s['border_box']
        c_fl.fill = s['fill_gris']

        # Fecha value E5
        c_fv = ws.cell(row=5, column=5)
        c_fv.value = fecha_fmt  # type: ignore
        c_fv.font = s['font_info_value']
        c_fv.alignment = s['align_center']
        c_fv.border = s['border_box']

        # N° label D6
        c_nl = ws.cell(row=6, column=4)
        c_nl.value = "N° de Pro-forma"  # type: ignore
        c_nl.font = s['font_info_label']
        c_nl.alignment = s['align_center']
        c_nl.border = s['border_box']
        c_nl.fill = s['fill_gris']

        # N° value E6
        c_nv = ws.cell(row=6, column=5)
        c_nv.value = cotizacion_data.get('numero_cotizacion', '')  # type: ignore
        c_nv.font = s['font_info_value']
        c_nv.alignment = s['align_center']
        c_nv.border = s['border_box']

        # Ensure we're past both company info and fecha/numero
        next_row = max(row, 8) + 1

        # Blank separator row
        ws.row_dimensions[next_row].height = 6
        next_row += 1

        return next_row

    # ══════════════════════════════════════════════════════════
    #  ENCABEZADO DE TABLA (fila azul)
    # ══════════════════════════════════════════════════════════

    def _escribir_encabezado_tabla(self, ws: Worksheet, row: int, estilos: Dict[str, Any]) -> int:
        s = estilos
        ws.row_dimensions[row].height = 24

        headers = ['IVA', 'CANT.', 'DESCRIPCIÓN', 'P. UNITARIO', 'TOTAL']
        for col_idx, header_text in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx)
            c.value = header_text  # type: ignore
            c.font = s['font_header']
            c.fill = s['fill_azul']
            c.alignment = s['align_center']
            c.border = Border(
                left=Side(style='thin', color=self.BLANCO),
                right=Side(style='thin', color=self.BLANCO),
                top=Side(style='thin', color=self.BLANCO),
                bottom=Side(style='thin', color=self.BLANCO),
            )

        return row + 1

    # ══════════════════════════════════════════════════════════
    #  FILAS DE CONCEPTOS (con separadores de grupo/ciudad)
    # ══════════════════════════════════════════════════════════

    def _escribir_conceptos(
        self, ws: Worksheet, start_row: int, detalles: List[Dict[str, Any]], estilos: Dict[str, Any]
    ) -> int:
        s = estilos
        row = start_row
        current_grupo = None

        if not detalles:
            # 5 empty rows
            for _ in range(5):
                self._fila_vacia(ws, row, s)
                row += 1
            return row

        for detalle in detalles:
            grupo = detalle.get('grupo', '') or ''

            # Group separator row
            if grupo and grupo != current_grupo:
                current_grupo = grupo
                ws.merge_cells(f'A{row}:E{row}')
                c_grupo = ws.cell(row=row, column=1)
                c_grupo.value = grupo  # type: ignore
                c_grupo.font = s['font_grupo']
                c_grupo.alignment = s['align_center']
                c_grupo.fill = s['fill_gris']
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = s['border_thin']
                ws.row_dimensions[row].height = 20
                row += 1

            # Data row
            ws.row_dimensions[row].height = 20
            cantidad = detalle.get('cantidad', 0)
            precio_unitario = detalle.get('precio_unitario', 0)

            # A: IVA = TOTAL × 1.16 (formula)
            c_iva = ws.cell(row=row, column=1)
            c_iva.value = f'=E{row}*1.16'  # type: ignore
            c_iva.font = s['font_data']
            c_iva.alignment = s['align_right']
            c_iva.border = s['border_thin']
            c_iva.number_format = '#,##0.00'

            # B: CANT
            c_cant = ws.cell(row=row, column=2)
            c_cant.value = int(cantidad) if cantidad == int(cantidad) else cantidad  # type: ignore
            c_cant.font = s['font_data']
            c_cant.alignment = s['align_center']
            c_cant.border = s['border_thin']

            # C: DESCRIPCIÓN
            c_desc = ws.cell(row=row, column=3)
            c_desc.value = detalle.get('descripcion', '')  # type: ignore
            c_desc.font = s['font_data']
            c_desc.alignment = s['align_center']
            c_desc.border = s['border_thin']

            # D: P. UNITARIO
            c_pu = ws.cell(row=row, column=4)
            c_pu.value = precio_unitario  # type: ignore
            c_pu.font = s['font_data']
            c_pu.alignment = s['align_right']
            c_pu.border = s['border_thin']
            c_pu.number_format = '#,##0.00'

            # E: TOTAL = CANT × P.U. (formula)
            c_total = ws.cell(row=row, column=5)
            c_total.value = f'=B{row}*D{row}'  # type: ignore
            c_total.font = s['font_data']
            c_total.alignment = s['align_right']
            c_total.border = s['border_thin']
            c_total.number_format = '#,##0.00'

            row += 1

        # Minimum 25 rows of table (fill empty rows with $0.00 formulas)
        data_count = sum(1 for d in detalles if d.get('cantidad'))
        group_count = len(set(d.get('grupo', '') for d in detalles if d.get('grupo', '')))
        total_rows = data_count + group_count
        min_rows = 25
        filas_extra = max(0, min_rows - total_rows)
        for _ in range(filas_extra):
            self._fila_vacia(ws, row, s)
            row += 1

        return row

    def _fila_vacia(self, ws: Worksheet, row: int, estilos: Dict[str, Any]) -> None:
        s = estilos
        ws.row_dimensions[row].height = 18
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.border = s['border_thin']

        # IVA = 0.00 formula (references TOTAL col)
        ws.cell(row=row, column=1).value = f'=E{row}*1.16'  # type: ignore
        ws.cell(row=row, column=1).number_format = '#,##0.00'
        ws.cell(row=row, column=1).font = s['font_data']
        ws.cell(row=row, column=1).alignment = s['align_right']

        # P.UNITARIO
        ws.cell(row=row, column=4).number_format = '#,##0.00'

        # TOTAL = 0.00
        ws.cell(row=row, column=5).value = 0  # type: ignore
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=5).font = s['font_data']
        ws.cell(row=row, column=5).alignment = s['align_right']

    # ══════════════════════════════════════════════════════════
    #  BLOQUE DE TOTALES
    # ══════════════════════════════════════════════════════════

    def _escribir_totales(
        self, ws: Worksheet, row: int, data_start: int, data_end: int,
        cotizacion_data: Dict[str, Any], estilos: Dict[str, Any]
    ) -> int:
        s = estilos
        last = data_end - 1

        row += 1  # space row

        descuento = cotizacion_data.get('descuento', 0) or 0
        envio = cotizacion_data.get('envio_delivery', 0) or 0

        # Helper to write a totals row: label in C:D merged, value in E
        def _totals_row(r: int, label: str, formula_or_value: Any,
                        label_font: Any = None, value_font: Any = None,
                        value_fill: Any = None, label_fill: Any = None,
                        number_fmt: str = '#,##0.00') -> None:
            ws.row_dimensions[r].height = 20
            ws.merge_cells(f'C{r}:D{r}')
            c_lbl = ws.cell(row=r, column=3)
            c_lbl.value = label  # type: ignore
            c_lbl.font = label_font or s['font_totales_label']
            c_lbl.alignment = Alignment(horizontal='right', vertical='center')
            if label_fill:
                c_lbl.fill = label_fill
                ws.cell(row=r, column=4).fill = label_fill

            c_val = ws.cell(row=r, column=5)
            c_val.value = formula_or_value  # type: ignore
            c_val.font = value_font or s['font_totales_value']
            c_val.alignment = Alignment(horizontal='right', vertical='center')
            c_val.number_format = number_fmt
            if value_fill:
                c_val.fill = value_fill
            c_val.border = Border(
                bottom=Side(style='thin', color='CCCCCC'),
            )

        # Total parcial
        _totals_row(row, 'Total parcial', f'=SUM(E{data_start}:E{last})')
        total_parcial_row = row
        row += 1

        # Descuento ($)
        _totals_row(row, 'Descuento ($)', descuento)
        descuento_row = row
        row += 1

        # NETO
        _totals_row(row, 'NETO', f'=E{total_parcial_row}-E{descuento_row}')
        neto_row = row
        row += 1

        # Impuesto (IVA)
        _totals_row(row, 'Impuesto (IVA)', f'=E{neto_row}*16%')
        iva_row = row
        row += 1

        # Envío Delivery
        _totals_row(row, 'Envío Delivery', envio)
        envio_row = row
        row += 1

        # Pagado (highlighted blue)
        _totals_row(
            row, 'Pagado',
            f'=E{neto_row}+E{iva_row}+E{envio_row}',
            label_font=s['font_pagado_label'],
            value_font=s['font_pagado_value'],
            label_fill=s['fill_azul'],
            value_fill=s['fill_azul'],
            number_fmt='#,##0.00',
        )
        # Also fill the merged cell D
        ws.cell(row=row, column=4).fill = s['fill_azul']
        # Bold border around pagado
        for col in range(3, 6):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='medium', color=self.AZUL_CORP),
                right=Side(style='medium', color=self.AZUL_CORP),
                top=Side(style='medium', color=self.AZUL_CORP),
                bottom=Side(style='medium', color=self.AZUL_CORP),
            )

        row += 2
        return row

    # ══════════════════════════════════════════════════════════
    #  TÉRMINOS Y CONDICIONES + PIE
    # ══════════════════════════════════════════════════════════

    def _escribir_terminos_y_pie(
        self, ws: Worksheet, row: int, cotizacion_data: Dict[str, Any],
        empresa_data: Dict[str, Any], estilos: Dict[str, Any]
    ) -> int:
        s = estilos

        # ── Términos y Condiciones (left side A:C) ──
        notas = cotizacion_data.get('notas', '')
        if notas:
            ws.merge_cells(f'A{row}:C{row}')
            c_t = ws.cell(row=row, column=1)
            c_t.value = "Términos y Condiciones:"  # type: ignore
            c_t.font = s['font_terms_title']
            c_t.alignment = s['align_left']
            row += 1

            for linea in notas.split('\n'):
                if linea.strip():
                    ws.merge_cells(f'A{row}:C{row}')
                    c = ws.cell(row=row, column=1)
                    c.value = linea.strip()  # type: ignore
                    c.font = s['font_terms']
                    c.alignment = s['align_left']
                    ws.row_dimensions[row].height = 14
                    row += 1

        row += 2

        # ── Footer ──
        ws.merge_cells(f'A{row}:E{row}')
        c = ws.cell(row=row, column=1)
        c.value = empresa_data.get('nombre', '')  # type: ignore
        c.font = s['font_footer']
        c.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        ws.merge_cells(f'A{row}:E{row}')
        c = ws.cell(row=row, column=1)
        c.value = "Quedo a sus ordenes"  # type: ignore
        c.font = s['font_footer_sub']
        c.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        ws.merge_cells(f'A{row}:E{row}')
        c = ws.cell(row=row, column=1)
        c.value = "Saludos!"  # type: ignore
        c.font = s['font_footer_sub']
        c.alignment = Alignment(horizontal='center', vertical='center')

        return row

    # ══════════════════════════════════════════════════════════
    #  MÉTODO PRINCIPAL
    # ══════════════════════════════════════════════════════════

    def generar_cotizacion(self, cotizacion_data: Dict[str, Any], empresa_data: Dict[str, Any]) -> str:
        numero = cotizacion_data.get('numero_cotizacion', 'SIN-NUMERO')
        filename = f"{str(numero).replace('/', '-')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore
        estilos = self._crear_estilos()

        # 1. Configurar hoja
        self._configurar_hoja(ws)

        # 2. Encabezado (logo, pro-forma, empresa, fecha/numero)
        row = self._escribir_encabezado(ws, empresa_data, cotizacion_data, estilos)

        # 3. Encabezado de tabla
        row = self._escribir_encabezado_tabla(ws, row, estilos)

        # 4. Conceptos
        data_start = row
        detalles = cotizacion_data.get('detalles', [])
        row = self._escribir_conceptos(ws, row, detalles, estilos)
        data_end = row

        # 5. Totales
        row = self._escribir_totales(ws, row, data_start, data_end, cotizacion_data, estilos)

        # 6. Términos y pie
        self._escribir_terminos_y_pie(ws, row, cotizacion_data, empresa_data, estilos)

        wb.save(filepath)
        return filepath
